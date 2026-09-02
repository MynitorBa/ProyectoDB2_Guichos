import re
from datetime import date, datetime, time, timezone, timedelta
from decimal import Decimal
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import Response
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from pydantic import BaseModel, ConfigDict, Field
from pymongo.database import Database
from sqlalchemy import func
from sqlalchemy.orm import Session

GT_TZ = timezone(timedelta(hours=-6))

from bson import ObjectId

from app.core.db_mongo import get_mongo_db
from app.core.db_mysql import get_db
from app.core.deps import get_admin_user
from app.models.categoria import Categoria
from app.models.inventario import Inventario
from app.models.pago import Pago
from app.models.pedido import Pedido
from app.models.pedido_vendedor import PedidoVendedor
from app.models.producto_referencia import ProductoReferencia
from app.models.producto_variante_referencia import ProductoVarianteReferencia
from app.models.producto_referencia_categoria import ProductoReferenciaCategoria
from app.models.usuario import Rol, Usuario, UsuarioRol
from app.models.vendedor import Vendedor
from app.models.oferta import Oferta, OfertaPrecioHistorial
from app.models.producto_imagen import ProductoImagen
from app.models.solicitud_catalogo import SolicitudCatalogoImagen
from app.schemas.producto import ProductoCreate, ProductoUpdate
from app.services import catalog_service
from app.services.product_history_service import reconstruir_estado, obtener_historial
from app.services.outbox_service import enqueue_outbox
from app.services.offer_service import (
    actualizar_precio_oferta,
    enqueue_primary_offer_projection,
    listar_ofertas_por_referencias,
    oferta_principal,
)
from app.services.offer_history_service import (
    historial_precios_diario,
    historial_operativo_unificado,
    reconstruir_ofertas_en_fecha,
    registrar_estado_oferta,
    registrar_saldo_inventario,
)
from app.core.time import utc_now
from app.services.image_service import read_valid_image
from app.services.category_attribute_service import (
    AttributeValidationError,
    validate_category_attributes,
)
from app.services.sku_service import generate_offer_sku, generate_product_sku
from app.services.variant_service import create_variant, list_variants, delete_variant, update_variant_attributes

router = APIRouter(prefix='/admin', tags=['Admin'])


class RolesUpdate(BaseModel):
    roles: list[str]


class StatusUpdate(BaseModel):
    estado: str

class VendorProfilePayload(BaseModel):
    nombre_comercial: str
    nit: str


class OfferCreate(BaseModel):
    model_config = ConfigDict(extra='forbid')
    vendedor_id: int
    producto_variante_id: int
    precio: Decimal = Field(gt=0)
    stock: int = Field(ge=0)


class OfferUpdate(BaseModel):
    precio: Decimal | None = Field(default=None, gt=0)
    stock: int | None = Field(default=None, ge=0)
    estado: str | None = None


class VariantCreate(BaseModel):
    model_config = ConfigDict(extra='forbid')
    atributos: dict


class VariantUpdate(BaseModel):
    model_config = ConfigDict(extra='forbid')
    atributos: dict


class AtributoEsquema(BaseModel):
    nombre: str
    etiqueta: str
    tipo: str = 'string'        # string | number | boolean
    requerido: bool = False
    placeholder: str | None = None


class CategoriaCreate(BaseModel):
    nombre: str
    slug: str
    descripcion: str | None = None
    padre_id: int | None = None
    sku_prefix: str | None = None
    atributos: list[AtributoEsquema] = []


class EsquemaUpdate(BaseModel):
    atributos: list[AtributoEsquema]
    categoria_nombre: str | None = None
    sku_prefix: str | None = None


# ── Gestión de usuarios ───────────────────────────────────────────────────────

# Lista todos los usuarios del sistema con paginación; solo accesible por administradores
@router.get('/users')
def list_users(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    _: Usuario = Depends(get_admin_user),
    db: Session = Depends(get_db),
):
    total = db.query(func.count(Usuario.id)).scalar()
    usuarios = (
        db.query(Usuario)
        .order_by(Usuario.fecha_alta.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    return {
        'items': [
            {
                'id': u.id,
                'nombre': u.nombre,
                'apellido': u.apellido,
                'email': u.email,
                'estado': u.estado,
                'roles': [r.nombre for r in u.roles],
                'fecha_alta': u.fecha_alta.isoformat() if u.fecha_alta else None,
            }
            for u in usuarios
        ],
        'total': total,
        'page': page,
        'page_size': page_size,
        'total_pages': max(1, -(-total // page_size)),
    }


# Cruza usuarios con rol 'vendedor' y sus perfiles de Vendedor para mostrar NIT y nombre comercial
@router.get('/vendors')
def list_vendors(
    _: Usuario = Depends(get_admin_user),
    db: Session = Depends(get_db),
):
    rol_vendedor = db.query(Rol).filter_by(nombre='vendedor').first()
    if not rol_vendedor:
        return []
    usuario_ids = [ur.usuario_id for ur in db.query(UsuarioRol).filter_by(rol_id=rol_vendedor.id).all()]
    if not usuario_ids:
        return []
    usuarios = db.query(Usuario).filter(Usuario.id.in_(usuario_ids)).all()
    vendedores = {v.usuario_id: v for v in db.query(Vendedor).filter(Vendedor.usuario_id.in_(usuario_ids)).all()}
    return [
        {
            'vendedor_id': vendedores[u.id].id if u.id in vendedores else None,
            'usuario_id': u.id,
            'nombre_completo': f'{u.nombre} {u.apellido}',
            'email': u.email,
            'nombre_comercial': vendedores[u.id].nombre_comercial if u.id in vendedores else None,
            'nit': vendedores[u.id].nit if u.id in vendedores else None,
            'estado_verificacion': vendedores[u.id].estado_verificacion if u.id in vendedores else None,
            'es_tiendaya': bool(vendedores[u.id].es_tiendaya) if u.id in vendedores else False,
        }
        for u in usuarios
    ]


# Asigna el flag es_tiendaya a un único vendedor; quita el flag de todos los demás (asignación exclusiva)
@router.patch('/vendors/{vendedor_id}/tiendaya')
def set_tiendaya_vendor(
    vendedor_id: int,
    _: Usuario = Depends(get_admin_user),
    db: Session = Depends(get_db),
):
    vendedor = db.get(Vendedor, vendedor_id)
    if not vendedor:
        raise HTTPException(404, 'Vendedor no encontrado.')
    # Si ya era TiendaYa, des-asignar (toggle off)
    if vendedor.es_tiendaya:
        vendedor.es_tiendaya = False
        db.commit()
        return {'vendedor_id': vendedor_id, 'es_tiendaya': False}
    # Quitar flag de cualquier otro vendedor que lo tenga
    db.query(Vendedor).filter(Vendedor.es_tiendaya == True).update({'es_tiendaya': False})  # noqa: E712
    vendedor.es_tiendaya = True
    db.commit()
    return {'vendedor_id': vendedor_id, 'es_tiendaya': True}


# Reemplaza completamente los roles de un usuario; impide que el admin se quite su propio rol
@router.patch('/users/{user_id}/roles')
def update_user_roles(
    user_id: int,
    payload: RolesUpdate,
    current_user: Usuario = Depends(get_admin_user),
    db: Session = Depends(get_db),
):
    if not payload.roles:
        raise HTTPException(400, 'El usuario debe tener al menos un rol.')

    if user_id == current_user.id and 'administrador' not in payload.roles:
        raise HTTPException(400, 'No puedes quitarte el rol de administrador a ti mismo.')

    usuario = db.get(Usuario, user_id)
    if not usuario:
        raise HTTPException(404, 'Usuario no encontrado.')

    roles_objs = db.query(Rol).filter(Rol.nombre.in_(payload.roles)).all()
    nombres_encontrados = {r.nombre for r in roles_objs}
    invalidos = set(payload.roles) - nombres_encontrados
    if invalidos:
        raise HTTPException(400, f'Roles no reconocidos: {", ".join(invalidos)}')

    db.query(UsuarioRol).filter(UsuarioRol.usuario_id == user_id).delete()
    db.flush()
    for rol in roles_objs:
        db.add(UsuarioRol(usuario_id=user_id, rol_id=rol.id))

    db.commit()
    db.refresh(usuario)
    return {'id': usuario.id, 'roles': [r.nombre for r in usuario.roles]}


# ── Estadísticas del catálogo ─────────────────────────────────────────────────

@router.get('/stats/catalog')
def stats_catalog(
    _: Usuario = Depends(get_admin_user),
    db: Database = Depends(get_mongo_db),
):
    return catalog_service.stats_catalogo(db)


# ── CRUD de productos (escribe en Mongo + genera eventos) ─────────────────────

@router.get('/products')
def listar_productos_admin(
    estado: str = Query('todos'),
    q: str | None = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    _: Usuario = Depends(get_admin_user),
    db: Database = Depends(get_mongo_db),
    mysql_db: Session = Depends(get_db),
):
    estados_validos = {'todos', 'activo', 'inactivo', 'descontinuado'}
    if estado not in estados_validos:
        raise HTTPException(422, f'Estado inválido: {estado}.')
    return catalog_service.listar_productos(
        db,
        mysql_db,
        q=q,
        page=page,
        page_size=page_size,
        orden='nombre_asc',
        estado=None if estado == 'todos' else estado,
    )

# Crea el documento en MongoDB y acto seguido crea ProductoReferencia, Oferta e Inventario en MySQL (con compensación si falla SQL)
@router.post('/products', status_code=201)
def crear_producto(
    payload: ProductoCreate,
    current_user: Usuario = Depends(get_admin_user),
    db: Database = Depends(get_mongo_db),
    mysql_db: Session = Depends(get_db),
):
    slugs = list(dict.fromkeys(payload.categoria_slugs))
    primary_slug = slugs[0]
    todas_categorias = mysql_db.query(Categoria).filter(
        Categoria.slug.in_(slugs), Categoria.activa.is_(True)
    ).all()
    by_slug = {category.slug: category for category in todas_categorias}
    invalidas = [slug for slug in slugs if slug not in by_slug]
    if invalidas:
        raise HTTPException(
            400,
            f'Categorías inexistentes o inactivas: {", ".join(invalidas)}',
        )
    categoria = by_slug[primary_slug]
    todas_categorias = [by_slug[slug] for slug in slugs]

    if not payload.vendedor_usuario_id:
        raise HTTPException(400, 'Selecciona el vendedor de la oferta inicial.')
    v_user = mysql_db.get(Usuario, payload.vendedor_usuario_id)
    v_rec = mysql_db.query(Vendedor).filter_by(
        usuario_id=payload.vendedor_usuario_id
    ).first()
    if not v_user or not v_rec:
        raise HTTPException(400, 'El usuario seleccionado no tiene perfil de vendedor.')

    # Verificar nombre duplicado antes de crear (case-insensitive)
    existing_doc = db.productos.find_one(
        {'nombre': {'$regex': f'^{re.escape(payload.nombre.strip())}$', '$options': 'i'}},
        {'_id': 1, 'nombre': 1},
    )
    if existing_doc:
        raise HTTPException(
            409,
            detail={
                'message': f'Ya existe un producto con el nombre "{existing_doc.get("nombre", payload.nombre)}".',
                'existing_id': str(existing_doc['_id']),
                'existing_nombre': existing_doc.get('nombre', payload.nombre),
            },
        )

    sku_prefix = categoria.sku_prefix or primary_slug[:3].upper()
    try:
        sku = generate_product_sku(db, sku_prefix)
        attributes = validate_category_attributes(db, slugs, payload.atributos)
    except (ValueError, AttributeValidationError) as exc:
        raise HTTPException(422, str(exc)) from exc
    v_nombre = v_rec.nombre_comercial
    v_mongo_id = v_rec.id
    cats_mongo = [
        {'slug': category.slug, 'nombre': category.nombre}
        for category in todas_categorias
    ]

    doc = {
        'sku': sku,
        'nombre': payload.nombre,
        'descripcion': payload.descripcion,
        'precio': float(payload.precio),
        'moneda': 'GTQ',
        'categoria': cats_mongo[0],
        'categorias': cats_mongo,
        'atributos': attributes,
        'imagenes': list(payload.imagenes),
        'vendedor_id': v_mongo_id,
        'vendedor_usuario_id': v_user.id,
        'vendedor_nombre': v_nombre,
        'resumen_resenas': {'promedio': 0.0, 'total': 0},
        'stock': payload.stock,
        'disponible': payload.stock > 0,
    }
    producto_mongo = catalog_service.crear_producto(
        db, doc, usuario_id=str(current_user.id)
    )

    # MySQL conserva solo identidad relacional, oferta, precio e inventario.
    try:
        ref = ProductoReferencia(
            producto_ref=producto_mongo['_id'], categoria_id=categoria.id
        )
        mysql_db.add(ref)
        mysql_db.flush()
        variant_registry, _ = create_variant(
            db,
            mysql_db,
            producto_ref=producto_mongo['_id'],
            attributes={},
            product_sku=sku,
            default=True,
        )
        for category in todas_categorias:
            mysql_db.add(ProductoReferenciaCategoria(
                producto_referencia_id=ref.id,
                categoria_id=category.id,
                es_principal=(category.slug == primary_slug),
            ))

        for orden, url in enumerate(payload.imagenes):
            try:
                img_id = int(url.rsplit('/', 1)[-1])
            except (ValueError, AttributeError):
                continue
            img_row = mysql_db.get(ProductoImagen, img_id)
            if img_row and img_row.producto_referencia_id is None:
                img_row.producto_referencia_id = ref.id
                img_row.orden = orden

        oferta = Oferta(
            producto_ref=producto_mongo['_id'],
            producto_variante_id=variant_registry.id,
            vendedor_id=v_rec.id,
            sku=sku,
            precio_actual=payload.precio,
            moneda='GTQ',
            estado='activa',
            version=1,
        )
        mysql_db.add(oferta)
        mysql_db.flush()
        mysql_db.add(OfertaPrecioHistorial(
            oferta_id=oferta.id,
            precio=payload.precio,
            moneda='GTQ',
            vigente_desde=utc_now(),
            cambiado_por=current_user.id,
            motivo='Precio inicial al crear la oferta',
        ))
        inventory = Inventario(
            oferta_id=oferta.id,
            cantidad_disponible=payload.stock,
            bodega='principal',
        )
        mysql_db.add(inventory)
        mysql_db.flush()
        registrar_estado_oferta(
            mysql_db, oferta=oferta, usuario_id=current_user.id,
            motivo='Estado inicial al crear la oferta', forzar=True,
        )
        registrar_saldo_inventario(
            mysql_db, inventario=inventory, usuario_id=current_user.id,
            motivo='Saldo inicial al crear la oferta', forzar=True,
        )
        enqueue_primary_offer_projection(mysql_db, producto_mongo['_id'], oferta.id)
        mysql_db.commit()
    except Exception:
        mysql_db.rollback()
        # Compensación: no dejar un documento comprable sin identidad/oferta SQL.
        db.productos.delete_one({'_id': ObjectId(producto_mongo['_id'])})
        db.producto_eventos.delete_many({'producto_id': producto_mongo['_id']})
        db.producto_variantes.delete_many({'producto_ref': producto_mongo['_id']})
        raise

    return producto_mongo


# Sube una imagen temporal a MySQL; queda huérfana hasta ser asignada a un producto
@router.post('/upload')
async def upload_image(
    file: UploadFile = File(...),
    current_user: Usuario = Depends(get_admin_user),
    mysql_db: Session = Depends(get_db),
):
    content, mime = await read_valid_image(file)
    img = ProductoImagen(
        datos=content, mime_type=mime, subida_por=current_user.id
    )
    mysql_db.add(img)
    mysql_db.commit()
    mysql_db.refresh(img)
    return {'url': f'/api/v1/products/images/{img.id}'}


@router.delete('/upload/{image_id}', status_code=204)
def delete_pending_admin_image(
    image_id: int,
    current_user: Usuario = Depends(get_admin_user),
    mysql_db: Session = Depends(get_db),
):
    """Descarta únicamente una carga temporal propia que aún no tiene dueño."""
    image = mysql_db.get(ProductoImagen, image_id)
    linked_request = mysql_db.query(SolicitudCatalogoImagen).filter_by(
        producto_imagen_id=image_id
    ).first()
    if not image:
        raise HTTPException(404, 'Imagen temporal no encontrada.')
    if (
        image.subida_por != current_user.id
        or image.producto_referencia_id is not None
        or linked_request is not None
    ):
        raise HTTPException(409, 'La imagen ya está asociada y no puede descartarse.')
    mysql_db.delete(image)
    mysql_db.commit()


# Actualización híbrida: los datos documentales van a MongoDB y precio/stock/vendedor se enrutan por outbox transaccional MySQL→Mongo
@router.put('/products/{producto_id}')
def actualizar_producto(
    producto_id: str,
    payload: ProductoUpdate,
    current_user: Usuario = Depends(get_admin_user),
    db: Database = Depends(get_mongo_db),
    mysql_db: Session = Depends(get_db),
):
    cambios = payload.model_dump(exclude_none=True)
    try:
        mongo_id = ObjectId(producto_id)
    except Exception as exc:
        raise HTTPException(404, 'Producto no encontrado.') from exc
    original_doc = db.productos.find_one({'_id': mongo_id})
    if not original_doc:
        raise HTTPException(404, 'Producto no encontrado.')

    if 'precio' in cambios:
        cambios['precio'] = float(cambios['precio'])

    nuevas_imagenes = cambios.pop('imagenes', None)

    nuevo_stock = None
    if 'stock' in cambios:
        nuevo_stock = cambios['stock']
        if nuevo_stock <= 0:
            cambios['disponible'] = False

    nuevo_vendedor_id_sql = None
    if 'vendedor_usuario_id' in cambios:
        v_uid = cambios.pop('vendedor_usuario_id')
        if v_uid:
            v_user = mysql_db.get(Usuario, v_uid)
            cambios['vendedor_id'] = str(v_uid)
            cambios['vendedor_usuario_id'] = v_uid
            cambios['vendedor_nombre'] = f'{v_user.nombre} {v_user.apellido}' if v_user else f'Usuario #{v_uid}'
            v_rec = mysql_db.query(Vendedor).filter_by(usuario_id=v_uid).first()
            if not v_rec:
                raise HTTPException(400, 'El usuario seleccionado no tiene perfil de vendedor.')
            nuevo_vendedor_id_sql = v_rec.id

    ofertas_producto = listar_ofertas_por_referencias(
        mysql_db, [producto_id], solo_activas=True
    ).get(producto_id, [])
    oferta_principal_data = oferta_principal(ofertas_producto)
    oferta = (
        mysql_db.get(Oferta, oferta_principal_data['oferta_id'])
        if oferta_principal_data else
        mysql_db.query(Oferta).filter_by(producto_ref=producto_id).order_by(Oferta.id).first()
    )
    cambios_operativos = {
        'precio', 'stock', 'estado', 'disponible', 'vendedor_usuario_id'
    }
    if any(key in payload.model_fields_set for key in cambios_operativos) and not oferta:
        raise HTTPException(
            status_code=409,
            detail='El producto documental no tiene una oferta MySQL asociada.',
        )

    vendedor_actualizado = bool(
        oferta
        and nuevo_vendedor_id_sql is not None
        and oferta.vendedor_id != nuevo_vendedor_id_sql
    )
    if vendedor_actualizado:
        oferta_duplicada = mysql_db.query(Oferta).filter(
            Oferta.producto_ref == producto_id,
            Oferta.vendedor_id == nuevo_vendedor_id_sql,
            Oferta.id != oferta.id,
            Oferta.estado != 'descontinuada',
        ).first()
        if oferta_duplicada:
            raise HTTPException(
                409,
                'El vendedor seleccionado ya tiene una oferta vigente para este producto.',
            )

    nuevos_slugs = cambios.pop('categoria_slugs', None)
    if 'atributos' in cambios:
        current_slugs = [
            category.get('slug') for category in original_doc.get('categorias', [])
            if category.get('slug')
        ]
        if not current_slugs and original_doc.get('categoria', {}).get('slug'):
            current_slugs = [original_doc['categoria']['slug']]
        try:
            cambios['atributos'] = validate_category_attributes(
                db, nuevos_slugs or current_slugs, cambios['atributos']
            )
        except AttributeValidationError as exc:
            raise HTTPException(422, str(exc)) from exc

    # Validar todas las referencias relacionales antes de tocar MongoDB. Así
    # una categoría o imagen inválida no deja una edición documental parcial.
    todas_cats = None
    cats_mongo = None
    if nuevos_slugs is not None:
        if not nuevos_slugs:
            raise HTTPException(422, 'Selecciona al menos una categoría.')
        todas_cats = mysql_db.query(Categoria).filter(
            Categoria.slug.in_(nuevos_slugs), Categoria.activa.is_(True)
        ).all()
        invalidas = sorted(set(nuevos_slugs) - {cat.slug for cat in todas_cats})
        if invalidas:
            raise HTTPException(
                400, f'Categorías inexistentes o inactivas: {", ".join(invalidas)}'
            )
        cats_mongo = [
            {'slug': category.slug, 'nombre': category.nombre}
            for category in sorted(
                todas_cats, key=lambda category: nuevos_slugs.index(category.slug)
            )
        ]

    ref_img = None
    current_imgs = []
    image_rows: dict[int, ProductoImagen] = {}
    new_image_ids: set[int] = set()
    if nuevas_imagenes is not None:
        ref_img = mysql_db.query(ProductoReferencia).filter_by(
            producto_ref=producto_id
        ).first()
        if not ref_img:
            raise HTTPException(409, 'El producto no tiene referencia SQL para sus imágenes.')
        current_imgs = mysql_db.query(ProductoImagen).filter_by(
            producto_referencia_id=ref_img.id
        ).all()
        for url in nuevas_imagenes:
            prefix = '/api/v1/products/images/'
            if not isinstance(url, str) or not url.startswith(prefix):
                continue  # Las URL externas heredadas siguen siendo válidas.
            try:
                image_id = int(url[len(prefix):])
            except ValueError as exc:
                raise HTTPException(400, f'URL de imagen inválida: {url}') from exc
            if image_id in new_image_ids:
                raise HTTPException(400, 'La misma imagen no puede repetirse.')
            image = mysql_db.get(ProductoImagen, image_id)
            if not image:
                raise HTTPException(400, f'La imagen {image_id} no existe.')
            if image.producto_referencia_id not in (None, ref_img.id):
                raise HTTPException(409, f'La imagen {image_id} pertenece a otro producto.')
            if (
                image.producto_referencia_id is None
                and image.subida_por != current_user.id
            ):
                raise HTTPException(403, f'La imagen temporal {image_id} pertenece a otro usuario.')
            new_image_ids.add(image_id)
            image_rows[image_id] = image

    # Mongo recibe solo datos documentales. Precio, stock y vendedor se
    # proyectan después desde el outbox transaccional MySQL.
    campos_transaccionales = {
        'precio', 'stock', 'disponible', 'vendedor_id', 'vendedor_usuario_id',
        'vendedor_nombre'
    }
    cambios_documentales = {
        key: value for key, value in cambios.items()
        if key not in campos_transaccionales
    }
    if nuevas_imagenes is not None:
        cambios_documentales['imagenes'] = nuevas_imagenes
    if cats_mongo:
        cambios_documentales.update({
            'categorias': cats_mongo,
            'categoria': cats_mongo[0],
        })
    producto = catalog_service.actualizar_producto(
        db, producto_id, cambios_documentales, usuario_id=str(current_user.id)
    )
    if not producto:
        raise HTTPException(status_code=404, detail='Producto no encontrado.')

    needs_commit = False

    if nuevas_imagenes is not None:
        for image in current_imgs:
            if image.id not in new_image_ids:
                mysql_db.delete(image)
        for order, url in enumerate(nuevas_imagenes):
            prefix = '/api/v1/products/images/'
            if isinstance(url, str) and url.startswith(prefix):
                image = image_rows[int(url[len(prefix):])]
                image.producto_referencia_id = ref_img.id
                image.orden = order
        needs_commit = True

    if nuevos_slugs is not None:
        if todas_cats:
            ref = mysql_db.query(ProductoReferencia).filter_by(producto_ref=producto_id).first()
            if ref:
                principal = next(cat for cat in todas_cats if cat.slug == nuevos_slugs[0])
                ref.categoria_id = principal.id
                mysql_db.query(ProductoReferenciaCategoria).filter_by(
                    producto_referencia_id=ref.id
                ).delete()
                for cat in todas_cats:
                    mysql_db.add(ProductoReferenciaCategoria(
                        producto_referencia_id=ref.id,
                        categoria_id=cat.id,
                        es_principal=(cat.slug == nuevos_slugs[0]),
                    ))
            needs_commit = True
    if oferta and 'precio' in cambios:
        if actualizar_precio_oferta(
            mysql_db,
            oferta=oferta,
            nuevo_precio=Decimal(str(cambios['precio'])),
            usuario_id=current_user.id,
            motivo='Actualización desde panel administrativo',
        ):
            needs_commit = True

    inv = (
        mysql_db.query(Inventario).filter_by(
            oferta_id=oferta.id, bodega='principal'
        ).first()
        if oferta else None
    )
    if nuevo_stock is not None and inv:
        inv.cantidad_disponible = nuevo_stock
        registrar_saldo_inventario(
            mysql_db, inventario=inv, usuario_id=current_user.id,
            motivo='Actualización de stock desde panel administrativo',
        )
        enqueue_outbox(
            mysql_db,
            tipo_evento='inventario.actualizado',
            agregado_tipo='oferta',
            agregado_id=oferta.id,
            producto_ref=producto_id,
            payload={
                'projection': {'stock': nuevo_stock, 'disponible': nuevo_stock > 0},
            },
        )
        needs_commit = True

    if vendedor_actualizado:
        oferta.vendedor_id = nuevo_vendedor_id_sql
        vendedor = mysql_db.get(Vendedor, nuevo_vendedor_id_sql)
        enqueue_outbox(
            mysql_db,
            tipo_evento='oferta.vendedor_actualizado',
            agregado_tipo='oferta',
            agregado_id=oferta.id,
            producto_ref=oferta.producto_ref,
            payload={'projection': {
                'vendedor_id': vendedor.id,
                'vendedor_usuario_id': vendedor.usuario_id,
                'vendedor_nombre': vendedor.nombre_comercial,
            }},
        )
        needs_commit = True

    if oferta and ('estado' in cambios or 'disponible' in cambios):
        if 'estado' in cambios:
            oferta.estado = {
                'activo': 'activa',
                'inactivo': 'pausada',
                'borrador': 'borrador',
                'descontinuado': 'descontinuada',
            }.get(cambios['estado'], oferta.estado)
        elif cambios.get('disponible') is False:
            oferta.estado = 'pausada'
        elif cambios.get('disponible') is True and oferta.estado == 'pausada':
            oferta.estado = 'activa'
        available = max(
            0,
            (inv.cantidad_disponible - inv.cantidad_reservada) if inv else 0,
        )
        enqueue_outbox(
            mysql_db,
            tipo_evento='oferta.estado_actualizado',
            agregado_tipo='oferta',
            agregado_id=oferta.id,
            producto_ref=oferta.producto_ref,
            payload={'projection': {
                'disponible': oferta.estado == 'activa' and available > 0
            }},
        )
        needs_commit = True

    if oferta and (
        vendedor_actualizado
        or 'estado' in cambios
        or 'disponible' in cambios
    ):
        registrar_estado_oferta(
            mysql_db, oferta=oferta, usuario_id=current_user.id,
            motivo='Actualización de oferta desde panel administrativo',
        )

    if needs_commit:
        try:
            mysql_db.commit()
        except Exception:
            mysql_db.rollback()
            # Compensa la actualización documental si fallara el commit SQL.
            db.productos.replace_one({'_id': mongo_id}, original_doc)
            raise

    return catalog_service.obtener_producto(db, producto_id, mysql_db)


# Borrado lógico: marca el producto como 'descontinuado' en MongoDB y descontinúa todas sus ofertas en MySQL
@router.delete('/products/{producto_id}', status_code=204)
def eliminar_producto(
    producto_id: str,
    current_user: Usuario = Depends(get_admin_user),
    db: Database = Depends(get_mongo_db),
    mysql_db: Session = Depends(get_db),
):
    eliminado = catalog_service.eliminar_producto(db, producto_id, usuario_id=str(current_user.id))
    if not eliminado:
        raise HTTPException(status_code=404, detail='Producto no encontrado.')
    ofertas = mysql_db.query(Oferta).filter_by(producto_ref=producto_id).all()
    for oferta in ofertas:
        oferta.estado = 'descontinuada'
        registrar_estado_oferta(
            mysql_db, oferta=oferta, usuario_id=current_user.id,
            motivo='Producto descontinuado desde panel administrativo',
        )
        enqueue_outbox(
            mysql_db,
            tipo_evento='oferta.estado_actualizado',
            agregado_tipo='oferta',
            agregado_id=oferta.id,
            producto_ref=oferta.producto_ref,
            payload={'projection': {'disponible': False}},
        )
    mysql_db.commit()


# ── Historial de eventos ───────────────────────────────────────────────────────

# Devuelve el historial unificado de un producto: eventos documentales de MongoDB + operativos de MySQL, filtrable por rango de fechas
@router.get('/products/{producto_id}/history')
def historial_producto(
    producto_id: str,
    desde: date | None = Query(None),
    hasta: date | None = Query(None),
    fuente: str = Query('todas'),
    _: Usuario = Depends(get_admin_user),
    db: Database = Depends(get_mongo_db),
    mysql_db: Session = Depends(get_db),
):
    if fuente not in {'todas', 'mongodb', 'mysql'}:
        raise HTTPException(422, 'Fuente inválida. Usa todas, mongodb o mysql.')
    if desde and hasta and desde > hasta:
        raise HTTPException(422, 'La fecha inicial no puede ser posterior a la fecha final.')
    desde_utc = (
        datetime.combine(desde, time.min, tzinfo=GT_TZ)
        .astimezone(timezone.utc).replace(tzinfo=None)
        if desde else None
    )
    hasta_utc = (
        datetime.combine(hasta, time.max, tzinfo=GT_TZ)
        .astimezone(timezone.utc).replace(tzinfo=None)
        if hasta else None
    )
    eventos = []
    if fuente in {'todas', 'mongodb'}:
        for event in obtener_historial(db, producto_id):
            # Precio, oferta e inventario tienen como fuente autoritativa MySQL.
            if event['tipo_evento'] in {'PRECIO_ACTUALIZADO', 'DISPONIBILIDAD_CAMBIADA'}:
                continue
            event['fuente'] = 'mongodb'
            event['entidad'] = 'producto'
            event_dt = datetime.fromisoformat(event['timestamp']).astimezone(
                timezone.utc
            ).replace(tzinfo=None)
            if ((desde_utc is None or event_dt >= desde_utc)
                    and (hasta_utc is None or event_dt <= hasta_utc)):
                eventos.append(event)
    if fuente in {'todas', 'mysql'}:
        eventos.extend(historial_operativo_unificado(
            mysql_db,
            producto_ref=producto_id,
            desde_utc=desde_utc,
            hasta_utc=hasta_utc,
        ))
    eventos.sort(key=lambda event: (event['timestamp'], event['_id']))
    if not eventos:
        raise HTTPException(status_code=404, detail='No hay historial para este producto.')
    return {
        'producto_id': producto_id,
        'eventos': eventos,
        'filtros': {
            'desde': desde.isoformat() if desde else None,
            'hasta': hasta.isoformat() if hasta else None,
            'fuente': fuente,
        },
    }


# Reconstruye el estado exacto del producto en una fecha dada haciendo replay de eventos (event sourcing)
@router.get('/products/{producto_id}/state-at')
def estado_en_fecha(
    producto_id: str,
    fecha: str = Query(..., description='Fecha en formato YYYY-MM-DD o YYYY-MM-DDTHH:MM:SS'),
    _: Usuario = Depends(get_admin_user),
    db: Database = Depends(get_mongo_db),
    mysql_db: Session = Depends(get_db),
):
    try:
        if 'T' in fecha:
            # Acepta HH:MM o HH:MM:SS — fromisoformat maneja ambos
            naive_gt = datetime.fromisoformat(fecha)
        else:
            naive_gt = datetime.strptime(fecha, '%Y-%m-%d').replace(
                hour=23, minute=59, second=59
            )
        # El usuario ingresa hora en horario de Guatemala → convertir a UTC para consultar Mongo
        fecha_utc = naive_gt.replace(tzinfo=GT_TZ).astimezone(timezone.utc).replace(tzinfo=None)
    except ValueError:
        raise HTTPException(
            status_code=422,
            detail='Formato de fecha inválido. Usa YYYY-MM-DD o YYYY-MM-DDTHH:MM'
        )

    estado = reconstruir_estado(db, producto_id, fecha_utc)
    if estado is None:
        raise HTTPException(
            status_code=404,
            detail=f'El producto no existía o no tiene eventos hasta {fecha}.'
        )
    estado['ofertas'] = reconstruir_ofertas_en_fecha(
        mysql_db, producto_ref=producto_id, instante_utc=fecha_utc
    )
    for legacy_field in (
        'precio', 'moneda', 'stock', 'oferta_id', 'vendedor_id',
        'vendedor_usuario_id', 'vendedor_nombre',
    ):
        estado.pop(legacy_field, None)
    estado['disponible'] = any(
        offer['disponible'] for offer in estado['ofertas']
    )
    estado['_fuentes'] = {
        'producto': 'MongoDB producto_eventos',
        'ofertas': 'MySQL historiales temporales',
    }
    return estado


# Historial de precios diario por oferta: útil para gráficas de tendencia de precios
@router.get('/products/{producto_id}/price-history')
def historial_precios_producto(
    producto_id: str,
    desde: date | None = Query(None),
    hasta: date | None = Query(None),
    _: Usuario = Depends(get_admin_user),
    mysql_db: Session = Depends(get_db),
):
    try:
        result = historial_precios_diario(
            mysql_db, producto_ref=producto_id, desde=desde, hasta=hasta
        )
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    if not result['ofertas']:
        raise HTTPException(
            status_code=404,
            detail='No hay historial de precios para las ofertas de este producto.',
        )
    return result


# ── Migración de stock ────────────────────────────────────────────────────────

@router.post('/migrate-stock', status_code=200)
def migrar_stock(
    _: Usuario = Depends(get_admin_user),
    db: Database = Depends(get_mongo_db),
    mysql_db: Session = Depends(get_db),
):
    """Encola una reconciliación de stock MySQL → MongoDB."""
    actualizados = 0
    rows = (
        mysql_db.query(Oferta, Inventario)
        .join(
            Inventario,
            (Inventario.oferta_id == Oferta.id)
            & (Inventario.bodega == 'principal'),
        )
        .all()
    )
    for oferta, inv in rows:
        if db.productos.find_one({'_id': ObjectId(oferta.producto_ref)}, {'_id': 1}):
            enqueue_outbox(
                mysql_db,
                tipo_evento='inventario.reconciliado',
                agregado_tipo='oferta',
                agregado_id=oferta.id,
                producto_ref=oferta.producto_ref,
                payload={'projection': {
                    'stock': inv.cantidad_disponible,
                    'disponible': (
                        oferta.estado == 'activa'
                        and inv.cantidad_disponible - inv.cantidad_reservada > 0
                    ),
                }},
            )
            actualizados += 1
    mysql_db.commit()
    return {'encolados': actualizados}


# ── Gestión de categorías ─────────────────────────────────────────────────────

@router.get('/categories')
def listar_categorias_admin(
    _: Usuario = Depends(get_admin_user),
    db: Session = Depends(get_db),
    mongo: Database = Depends(get_mongo_db),
):
    """Lista categorías de MySQL enriquecidas con su esquema de MongoDB."""
    categorias = db.query(Categoria).order_by(Categoria.orden, Categoria.nombre).all()
    esquemas = {
        e['categoria_slug']: e.get('atributos', [])
        for e in mongo.categoria_esquemas.find({}, {'categoria_slug': 1, 'atributos': 1})
    }
    return [
        {
            'id': c.id,
            'nombre': c.nombre,
            'slug': c.slug,
            'sku_prefix': c.sku_prefix,
            'descripcion': c.descripcion,
            'padre_id': c.categoria_padre_id,
            'activa': c.activa,
            'atributos': esquemas.get(c.slug, []),  # schema desde Mongo
        }
        for c in categorias
    ]


@router.post('/categories', status_code=201)
def crear_categoria(
    payload: CategoriaCreate,
    _: Usuario = Depends(get_admin_user),
    db: Session = Depends(get_db),
    mongo: Database = Depends(get_mongo_db),
):
    """
    Crea la categoría en MySQL (nombre, slug, jerarquía) y su definición
    de campos en MongoDB (categoria_esquemas). Cada BD almacena lo que le
    corresponde según su naturaleza.
    """
    if db.query(Categoria).filter_by(slug=payload.slug).first():
        raise HTTPException(400, f'Ya existe una categoría con slug "{payload.slug}".')

    # Verificar nombre duplicado (case-insensitive)
    existing_cat = db.query(Categoria).filter(
        Categoria.nombre.ilike(payload.nombre.strip())
    ).first()
    if existing_cat:
        raise HTTPException(
            409,
            detail={
                'message': f'Ya existe una categoría con el nombre "{existing_cat.nombre}".',
                'existing_slug': existing_cat.slug,
                'existing_nombre': existing_cat.nombre,
            },
        )

    if payload.sku_prefix:
        prefix = payload.sku_prefix[:3].upper()
        conflict = db.query(Categoria).filter_by(sku_prefix=prefix).first()
        if conflict:
            raise HTTPException(400, f'El prefijo "{prefix}" ya está en uso por la categoría "{conflict.nombre}".')

    # 1. Guardar estructura relacional en MySQL
    cat = Categoria(
        nombre=payload.nombre,
        slug=payload.slug,
        sku_prefix=payload.sku_prefix[:3].upper() if payload.sku_prefix else None,
        descripcion=payload.descripcion,
        categoria_padre_id=payload.padre_id,
        activa=True,
    )
    db.add(cat)
    db.commit()
    db.refresh(cat)

    # 2. Guardar definición de atributos en MongoDB (upsert por slug)
    mongo.categoria_esquemas.update_one(
        {'categoria_slug': payload.slug},
        {'$set': {
            'categoria_slug': payload.slug,
            'categoria_nombre': payload.nombre,
            'atributos': [a.model_dump() for a in payload.atributos],
        }},
        upsert=True,
    )

    return {'id': cat.id, 'slug': cat.slug, 'nombre': cat.nombre}


@router.put('/categories/{slug}/schema')
def actualizar_esquema(
    slug: str,
    payload: EsquemaUpdate,
    _: Usuario = Depends(get_admin_user),
    db: Session = Depends(get_db),
    mongo: Database = Depends(get_mongo_db),
):
    """Actualiza los campos del esquema en MongoDB sin tocar MySQL."""
    cat = db.query(Categoria).filter_by(slug=slug).first()
    if not cat:
        raise HTTPException(404, 'Categoría no encontrada.')

    if payload.sku_prefix is not None:
        new_prefix = payload.sku_prefix[:3].upper() if payload.sku_prefix else None
        if new_prefix:
            conflict = db.query(Categoria).filter(
                Categoria.sku_prefix == new_prefix,
                Categoria.id != cat.id,
            ).first()
            if conflict:
                raise HTTPException(400, f'El prefijo "{new_prefix}" ya está en uso por la categoría "{conflict.nombre}".')
        cat.sku_prefix = new_prefix
        db.commit()

    mongo.categoria_esquemas.update_one(
        {'categoria_slug': slug},
        {'$set': {
            'categoria_nombre': payload.categoria_nombre or cat.nombre,
            'atributos': [a.model_dump() for a in payload.atributos],
        }},
        upsert=True,
    )
    return {'slug': slug, 'atributos': len(payload.atributos), 'sku_prefix': cat.sku_prefix}


@router.delete('/categories/{slug}', status_code=204)
def eliminar_categoria(
    slug: str,
    _: Usuario = Depends(get_admin_user),
    db: Session = Depends(get_db),
    mongo: Database = Depends(get_mongo_db),
):
    """Elimina una categoría sin uso o desactiva una categoría referenciada."""
    cat = db.query(Categoria).filter_by(slug=slug).first()
    if not cat:
        raise HTTPException(404, 'Categoría no encontrada.')

    referencias = db.query(ProductoReferencia).filter_by(
        categoria_id=cat.id
    ).count()
    if referencias or cat.hijos:
        cat.activa = False
        db.commit()
        return

    db.delete(cat)
    db.commit()
    mongo.categoria_esquemas.delete_one({'categoria_slug': slug})


# ── Panel de ventas ───────────────────────────────────────────────────────────

# Agrega KPIs de ventas: ingresos, pedidos por estado, serie de 30 días y top 5 vendedores
@router.get('/sales/stats')
def sales_stats(
    _: Usuario = Depends(get_admin_user),
    db: Session = Depends(get_db),
):
    total_pedidos = db.query(func.count(Pedido.id)).scalar() or 0

    agg = db.query(
        func.sum(Pedido.total),
        func.avg(Pedido.total),
    ).filter(Pedido.estado.notin_(['cancelado', 'reembolsado'])).first()
    total_ingresos = float(agg[0] or 0)
    promedio_pedido = float(agg[1] or 0)

    pendientes = db.query(func.count(Pedido.id)).filter(Pedido.estado == 'pendiente').scalar() or 0

    por_estado = [
        {'estado': e, 'cantidad': c}
        for e, c in db.query(Pedido.estado, func.count(Pedido.id)).group_by(Pedido.estado).all()
    ]

    hoy_gt = datetime.now(GT_TZ).date()
    hace_30 = hoy_gt - timedelta(days=29)
    cutoff = datetime(hace_30.year, hace_30.month, hace_30.day)

    dia_rows = db.query(
        func.date(Pedido.fecha_creacion).label('dia'),
        func.sum(Pedido.total).label('total'),
    ).filter(
        Pedido.fecha_creacion >= cutoff,
        Pedido.estado.notin_(['cancelado', 'reembolsado']),
    ).group_by('dia').all()

    dias_map = {str(r.dia): float(r.total) for r in dia_rows}
    ingresos_por_dia = [
        {
            'fecha': str(hace_30 + timedelta(days=i)),
            'total': dias_map.get(str(hace_30 + timedelta(days=i)), 0.0),
        }
        for i in range(30)
    ]

    top_rows = (
        db.query(
            Vendedor.nombre_comercial,
            func.sum(PedidoVendedor.subtotal).label('ingresos'),
            func.count(PedidoVendedor.id).label('pedidos'),
        )
        .join(PedidoVendedor, PedidoVendedor.vendedor_id == Vendedor.id)
        .join(Pedido, Pedido.id == PedidoVendedor.pedido_id)
        .filter(
            Pedido.estado.notin_(['cancelado', 'reembolsado']),
            PedidoVendedor.estado.notin_(['cancelado', 'reembolsado']),
        )
        .group_by(Vendedor.id, Vendedor.nombre_comercial)
        .order_by(func.sum(PedidoVendedor.subtotal).desc())
        .limit(5)
        .all()
    )
    top_vendedores = [
        {'nombre': r.nombre_comercial, 'ingresos': float(r.ingresos), 'pedidos': r.pedidos}
        for r in top_rows
    ]

    return {
        'total_pedidos': total_pedidos,
        'total_ingresos': total_ingresos,
        'promedio_pedido': promedio_pedido,
        'pendientes': pendientes,
        'por_estado': por_estado,
        'ingresos_por_dia': ingresos_por_dia,
        'top_vendedores': top_vendedores,
    }


@router.get('/sales')
def list_sales(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    _: Usuario = Depends(get_admin_user),
    db: Session = Depends(get_db),
):
    total = db.query(func.count(Pedido.id)).scalar() or 0
    pedidos = (
        db.query(Pedido)
        .order_by(Pedido.fecha_creacion.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )

    items = []
    for p in pedidos:
        u = p.usuario
        pago = p.pagos[0] if p.pagos else None
        items.append({
            'id': p.id,
            'fecha': p.fecha_creacion.replace(tzinfo=timezone.utc).astimezone(GT_TZ).isoformat(),
            'estado': p.estado,
            'subtotal': float(p.subtotal),
            'impuestos': float(p.impuestos),
            'total': float(p.total),
            'comprador': {
                'id': u.id,
                'nombre': f'{u.nombre} {u.apellido}',
                'email': u.email,
            } if u else None,
            'pago': {
                'metodo': pago.metodo.nombre if pago and pago.metodo else None,
                'estado': pago.estado,
                'referencia': pago.referencia_transaccion,
                'monto': float(pago.monto),
            } if pago else None,
            'lineas': [
                {
                    'producto_nombre': l.producto_nombre,
                    'precio_unitario': float(l.precio_unitario),
                    'cantidad': l.cantidad,
                    'subtotal_linea': float(l.subtotal_linea),
                    'vendedor': l.vendedor_nombre_snapshot,
                }
                for l in p.lineas
            ],
        })

    return {
        'items': items,
        'total': total,
        'page': page,
        'page_size': page_size,
        'total_pages': max(1, -(-total // page_size)),
    }


# Exporta todos los pedidos a Excel (.xlsx) con una fila por línea de producto; útil para reportes contables
@router.get('/sales/export')
def export_sales_excel(
    _: Usuario = Depends(get_admin_user),
    db: Session = Depends(get_db),
):
    pedidos = db.query(Pedido).order_by(Pedido.fecha_creacion.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Ventas'

    hdr_fill = PatternFill('solid', fgColor='2563EB')
    hdr_font = Font(color='FFFFFF', bold=True, size=10)
    hdr_align = Alignment(horizontal='center', vertical='center')

    COLS = [
        'Pedido #', 'Fecha (GT)', 'Estado', 'Comprador', 'Email',
        'Producto', 'Vendedor', 'Precio Unit.', 'Cantidad', 'Subtotal línea',
        'Impuestos', 'Total pedido', 'Método pago', 'Estado pago',
    ]
    for col, h in enumerate(COLS, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = hdr_align
    ws.row_dimensions[1].height = 18

    for p in pedidos:
        u = p.usuario
        pago = p.pagos[0] if p.pagos else None
        fecha_gt = p.fecha_creacion.replace(tzinfo=timezone.utc).astimezone(GT_TZ).strftime('%Y-%m-%d %H:%M')
        for l in p.lineas:
            ws.append([
                p.id,
                fecha_gt,
                p.estado,
                f'{u.nombre} {u.apellido}' if u else '—',
                u.email if u else '—',
                l.producto_nombre,
                l.vendedor_nombre_snapshot or '—',
                float(l.precio_unitario),
                l.cantidad,
                float(l.subtotal_linea),
                float(p.impuestos),
                float(p.total),
                pago.metodo.nombre if pago and pago.metodo else '—',
                pago.estado if pago else '—',
            ])

    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 45)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return Response(
        content=buf.read(),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename="ventas-TiendaYa.xlsx"'},
    )


# ── Pedidos (admin) ───────────────────────────────────────────────────────────

ESTADOS_VALIDOS = set(Pedido.__table__.c['estado'].type.enums)

@router.get('/orders')
def list_admin_orders(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    estado: str | None = Query(None),
    _: Usuario = Depends(get_admin_user),
    db: Session = Depends(get_db),
):
    q = db.query(Pedido)
    if estado:
        q = q.filter(Pedido.estado == estado)
    total = q.count()
    pedidos = (
        q.order_by(Pedido.fecha_creacion.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    return {
        'items': [
            {
                'id': p.id,
                'fecha': p.fecha_creacion.replace(tzinfo=timezone.utc).astimezone(GT_TZ).isoformat(),
                'estado': p.estado,
                'total': float(p.total),
                'comprador': {
                    'nombre': f'{p.usuario.nombre} {p.usuario.apellido}' if p.usuario else '—',
                    'email': p.usuario.email if p.usuario else '—',
                },
            }
            for p in pedidos
        ],
        'total': total,
        'page': page,
        'page_size': page_size,
        'total_pages': max(1, -(-total // page_size)),
    }


# Permite al admin cambiar el estado de cualquier pedido (incluyendo 'cancelado' y 'reembolsado')
@router.patch('/orders/{pedido_id}/status')
def update_admin_order_status(
    pedido_id: int,
    payload: StatusUpdate,
    _: Usuario = Depends(get_admin_user),
    db: Session = Depends(get_db),
):
    if payload.estado not in ESTADOS_VALIDOS:
        raise HTTPException(400, 'Estado no válido.')
    pedido = db.get(Pedido, pedido_id)
    if not pedido:
        raise HTTPException(404, 'Pedido no encontrado.')
    pedido.estado = payload.estado
    db.commit()
    return {'id': pedido_id, 'estado': payload.estado}


# ── Gestión de ofertas por producto ──────────────────────────────────────────

@router.get('/products/{producto_ref}/variants')
def get_product_variants(
    producto_ref: str,
    _: Usuario = Depends(get_admin_user),
    db: Session = Depends(get_db),
    mongo: Database = Depends(get_mongo_db),
):
    return list_variants(mongo, db, producto_ref)


@router.post('/products/{producto_ref}/variants', status_code=201)
def add_product_variant(
    producto_ref: str,
    payload: VariantCreate,
    _: Usuario = Depends(get_admin_user),
    db: Session = Depends(get_db),
    mongo: Database = Depends(get_mongo_db),
):
    try:
        product = mongo.productos.find_one(
            {'_id': ObjectId(producto_ref)}, {'sku': 1}
        )
    except Exception:
        product = None
    if not product:
        raise HTTPException(404, 'Producto no encontrado.')
    try:
        registry, _ = create_variant(
            mongo,
            db,
            producto_ref=producto_ref,
            attributes=payload.atributos,
            product_sku=product.get('sku', producto_ref[:8]),
        )
        db.commit()
    except ValueError as exc:
        db.rollback()
        raise HTTPException(422, str(exc)) from exc
    except Exception:
        db.rollback()
        raise
    return next(
        variant for variant in list_variants(mongo, db, producto_ref)
        if variant['variante_id'] == registry.id
    )

@router.patch('/variants/{variante_id}')
def patch_variant(
    variante_id: int,
    payload: VariantUpdate,
    _: Usuario = Depends(get_admin_user),
    db: Session = Depends(get_db),
    mongo: Database = Depends(get_mongo_db),
):
    try:
        result = update_variant_attributes(mongo, db, variante_id, payload.atributos)
        db.commit()
        return result
    except ValueError as e:
        db.rollback()
        raise HTTPException(400, str(e))


@router.delete('/variants/{variante_id}', status_code=204)
def delete_variant_endpoint(
    variante_id: int,
    _: Usuario = Depends(get_admin_user),
    db: Session = Depends(get_db),
    mongo: Database = Depends(get_mongo_db),
):
    try:
        delete_variant(mongo, db, variante_id)
        db.commit()
    except ValueError as e:
        db.rollback()
        raise HTTPException(400, str(e))


@router.get('/products/{producto_ref}/offers')
def list_product_offers(
    producto_ref: str,
    _: Usuario = Depends(get_admin_user),
    db: Session = Depends(get_db),
    mongo: Database = Depends(get_mongo_db),
):
    variants = {
        variant['variante_id']: variant
        for variant in list_variants(mongo, db, producto_ref)
    }
    rows = (
        db.query(Oferta, Vendedor)
        .join(Vendedor, Vendedor.id == Oferta.vendedor_id)
        .filter(Oferta.producto_ref == producto_ref)
        .order_by(Oferta.id)
        .all()
    )
    result = []
    for oferta, vendedor in rows:
        inv = db.query(Inventario).filter_by(oferta_id=oferta.id, bodega='principal').first()
        stock_disp = inv.cantidad_disponible if inv else 0
        stock_res = inv.cantidad_reservada if inv else 0
        result.append({
            'oferta_id': oferta.id,
            'vendedor_id': oferta.vendedor_id,
            'vendedor_nombre': vendedor.nombre_comercial,
            'producto_variante_id': oferta.producto_variante_id,
            'variante_ref': variants.get(
                oferta.producto_variante_id, {}
            ).get('variante_ref'),
            'variante_atributos': variants.get(
                oferta.producto_variante_id, {}
            ).get('atributos', {}),
            'sku': oferta.sku,
            'precio': float(oferta.precio_actual),
            'moneda': oferta.moneda,
            'estado': oferta.estado,
            'stock': max(0, stock_disp - stock_res),
            'stock_disponible': stock_disp,
            'version': oferta.version,
        })
    return result


# Agrega un vendedor adicional a un producto existente creando una nueva Oferta con su propio SKU e inventario
@router.post('/products/{producto_ref}/offers', status_code=201)
def add_product_offer(
    producto_ref: str,
    payload: OfferCreate,
    current_user: Usuario = Depends(get_admin_user),
    db: Session = Depends(get_db),
    mongo: Database = Depends(get_mongo_db),
):
    try:
        doc = mongo.productos.find_one({'_id': ObjectId(producto_ref)})
    except Exception:
        doc = None
    if not doc:
        raise HTTPException(404, 'Producto no encontrado.')

    vendedor = db.query(Vendedor).filter_by(id=payload.vendedor_id).first()
    if not vendedor:
        raise HTTPException(404, 'Vendedor no encontrado.')

    product_reference = db.query(ProductoReferencia).filter_by(
        producto_ref=producto_ref
    ).first()
    variant = db.get(ProductoVarianteReferencia, payload.producto_variante_id)
    if not product_reference or not variant or (
        variant.producto_referencia_id != product_reference.id
    ):
        raise HTTPException(400, 'La variante no pertenece al producto seleccionado.')
    variant_doc = mongo.producto_variantes.find_one(
        {'_id': ObjectId(variant.variante_ref)}
    )
    if not variant_doc or variant_doc.get('estado') != 'activa':
        raise HTTPException(400, 'La variante seleccionada no está activa.')

    existing = (
        db.query(Oferta)
        .filter_by(
            producto_variante_id=variant.id,
            vendedor_id=payload.vendedor_id,
        )
        .first()
    )
    if existing:
        raise HTTPException(
            400,
            'Este vendedor ya tiene una oferta para la variante; '
            'actualiza o reactiva la existente.',
        )

    try:
        sku = generate_offer_sku(
            db,
            product_sku=variant_doc.get(
                'sku_catalogo', doc.get('sku', producto_ref[:8])
            ),
            vendor_id=vendedor.id,
        )
    except ValueError as exc:
        raise HTTPException(409, str(exc)) from exc

    oferta = Oferta(
        producto_ref=producto_ref,
        producto_variante_id=variant.id,
        vendedor_id=payload.vendedor_id,
        sku=sku,
        precio_actual=payload.precio,
        moneda='GTQ',
        estado='activa',
        version=1,
    )
    db.add(oferta)
    db.flush()

    db.add(OfertaPrecioHistorial(
        oferta_id=oferta.id,
        precio=payload.precio,
        moneda='GTQ',
        vigente_desde=utc_now(),
        cambiado_por=current_user.id,
        motivo='Precio inicial de oferta adicional',
    ))
    inventory = Inventario(
        oferta_id=oferta.id,
        cantidad_disponible=payload.stock,
        bodega='principal',
    )
    db.add(inventory)
    db.flush()
    registrar_estado_oferta(
        db, oferta=oferta, usuario_id=current_user.id,
        motivo='Estado inicial de oferta adicional', forzar=True,
    )
    registrar_saldo_inventario(
        db, inventario=inventory, usuario_id=current_user.id,
        motivo='Saldo inicial de oferta adicional', forzar=True,
    )
    enqueue_primary_offer_projection(db, producto_ref, oferta.id)
    db.commit()

    return {
        'oferta_id': oferta.id,
        'vendedor_id': oferta.vendedor_id,
        'vendedor_nombre': vendedor.nombre_comercial,
        'producto_variante_id': variant.id,
        'variante_atributos': variant_doc.get('atributos', {}),
        'sku': oferta.sku,
        'precio': float(oferta.precio_actual),
        'estado': oferta.estado,
        'stock': payload.stock,
    }


# Actualiza precio, stock o estado de una oferta individual; encola proyección al finalizar
@router.patch('/offers/{oferta_id}')
def update_offer(
    oferta_id: int,
    payload: OfferUpdate,
    current_user: Usuario = Depends(get_admin_user),
    db: Session = Depends(get_db),
):
    oferta = db.get(Oferta, oferta_id)
    if not oferta:
        raise HTTPException(404, 'Oferta no encontrada.')

    if payload.precio is not None:
        actualizar_precio_oferta(
            db,
            oferta=oferta,
            nuevo_precio=payload.precio,
            usuario_id=current_user.id,
            motivo='Actualización desde panel admin',
            enqueue_projection=False,
        )

    if payload.stock is not None:
        inv = db.query(Inventario).filter_by(oferta_id=oferta_id, bodega='principal').first()
        if inv:
            inv.cantidad_disponible = payload.stock
        else:
            inv = Inventario(
                oferta_id=oferta_id, cantidad_disponible=payload.stock,
                bodega='principal'
            )
            db.add(inv)
        db.flush()
        registrar_saldo_inventario(
            db, inventario=inv, usuario_id=current_user.id,
            motivo='Actualización de stock desde panel admin',
        )

    if payload.estado is not None:
        estados_validos = {'activa', 'pausada', 'descontinuada', 'borrador'}
        if payload.estado not in estados_validos:
            raise HTTPException(400, f'Estado no válido. Opciones: {", ".join(estados_validos)}')
        oferta.estado = payload.estado

    if payload.estado is not None:
        registrar_estado_oferta(
            db, oferta=oferta, usuario_id=current_user.id,
            motivo='Actualización de estado desde panel admin',
        )

    db.flush()
    enqueue_primary_offer_projection(db, oferta.producto_ref, oferta.id)
    db.commit()
    return {'oferta_id': oferta_id, 'updated': True}


# ── Perfil de vendedor ────────────────────────────────────────────────────────

# Crea o actualiza el perfil de vendedor (nombre comercial + NIT) de un usuario existente
@router.post('/users/{user_id}/vendor-profile')
def set_vendor_profile(
    user_id: int,
    payload: VendorProfilePayload,
    _: Usuario = Depends(get_admin_user),
    db: Session = Depends(get_db),
):
    user = db.get(Usuario, user_id)
    if not user:
        raise HTTPException(404, 'Usuario no encontrado.')
    v = db.query(Vendedor).filter_by(usuario_id=user_id).first()
    if v:
        v.nombre_comercial = payload.nombre_comercial.strip()
        v.nit = payload.nit.strip()
    else:
        v = Vendedor(
            usuario_id=user_id,
            nombre_comercial=payload.nombre_comercial.strip(),
            nit=payload.nit.strip(),
        )
        db.add(v)
    try:
        db.commit()
    except Exception:
        db.rollback()
        raise HTTPException(400, 'El NIT ya está en uso por otro vendedor.')
    return {'usuario_id': user_id, 'nombre_comercial': v.nombre_comercial}
